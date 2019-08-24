import cognitive_face as CF
from global_variables import personGroupId
import os, urllib 
import sqlite3
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "D:/cam/reports.xlsx")
sheet = wb['Cse15']

def getDateColumn():
	for i in range(1, len(sheet['A']) + 1):
		col = get_column_letter(i)
		if sheet['%s%s'% (col,'1')].value == currentDate:
			return col
		
Key = '63032336225341b39b2ac1728b7864b7'
CF.Key.set(Key)
BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'  # Replace with your regional Base URL
CF.BaseUrl.set(BASE_URL)

connect = connect = sqlite3.connect(r"D:/cam/Face-DataBase")
c = connect.cursor()

attend = [0 for i in range(600)]	


directory = 'D:/cam/Cropped_faces'
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		imgurl = directory+"/"+ filename;  
		res = CF.face.detect(imgurl)
		if len(res) != 1:
			print ("No face detected.")
			continue
			
		faceIds = []
		for face in res:
			faceIds.append(face['faceId'])
		res = CF.face.identify(faceIds, personGroupId)
		print(filename)
		print( res)
		for face  in res:
			if not face['candidates']:
				print("Unknown")
			else:
				personId = face['candidates'][0]['personId']
				c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
				row = c.fetchone()
                #print(row)
				attend[int(row[0])] = 1;
				print( row[1] + " recognized")
for row in range(3, len(sheet['A']) + 1):
	rn = sheet['A%s'% row].value
	if rn is not None:
		rn = rn[-3:];
		if attend[int(rn)]!=0:
			col = getDateColumn();
			sheet["%s%s" % (col, str(row))]= 1
wb.save(filename = "D:/cam/reports.xlsx")	

#os.system('D:/Autoattendance-Cognitive-master/reports.xlsx')
#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res
